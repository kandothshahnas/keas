import openpyxl
from openpyxl import load_workbook


def get_row_count(file, sheetname):
    workbook = load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def get_column_count(file, sheetname):
    #workbook = openpyxl.load_workbook(file)
    #sheet = workbook.get_sheet_by_name(sheetname)
    workbook = load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def read_data(file, sheetname, rownum, columnno):
    #workbook = openpyxl.load_workbook(file)
    #sheet = workbook.get_sheet_by_name(sheetname)
    workbook = load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=columnno).value


def write_data(file, sheetname, rownum, columnno, data):
    #workbook = openpyxl.load_workbook(file)
    #sheet = workbook.get_sheet_by_name(sheetname)
    workbook = load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)
